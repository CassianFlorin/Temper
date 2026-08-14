"""优化 pass:查表去虚化 + IFERROR 剥除。

原理(见 docs/xlsx-anatomy.md):轴上 VLOOKUP/XLOOKUP 的查找键
($R25 技能名、S25:AL25 buff 名)与查找列(武学奇术!A:A、增益!A:A)
全部是静态字面量,与用户输入无关。因此行匹配可以在生成期完成,
运行时的线性扫描折叠为直接下标引用;去虚化后子树若不再含可抛错的
动态 VLOOKUP,包裹它的 IFERROR 一并剥掉。

不可折叠的边界由「输入格集合」显式给出:输入格及依赖它的一切保持
动态。默认输入区取 期望/RD 表 A1:I26 的字面量格(面板、三率、开关、
心法、怪物参数、战斗时间;见 anatomy 第 2.1/6 节 —— 轴区从 K 列起、
注释区从 27 行起,均不落入)。流派特有的武器格 F4:G7 恰在输入区内,
其对应的武器类型 VLOOKUP 自动保持动态,正确性不依赖启发式精确。

本 pass 不做「常量格折叠为缓存值」:缓存值带 15 位十进制截断,折叠
会让优化前后产生位级差异,失去逐位等价这条最强的验证手段。

作为脚本运行:python optimize_js.py <xlsx> [输出.js]
"""

from __future__ import annotations

import sys
from pathlib import Path
from typing import Optional

from openpyxl.utils import column_index_from_string, get_column_letter

from formula_parser import Bool, Call, Node, Num, Ref, Str
from emit_js import Emitter
from workbook_graph import WorkbookGraph, load_graph, resolve_ref

_EMPTY = object()  # 静态已知为空格(区别于「不可静态求值」的 None)

INPUT_MAX_COL = 9   # A..I
INPUT_MAX_ROW = 26


def default_inputs(g: WorkbookGraph) -> set[tuple[str, str]]:
    """期望/RD 表 A1:I26 内的字面量格。"""
    picked = set()
    for (sheet, coord), _v in g.values.items():
        if sheet not in ("期望", "RD", "DIY"):
            continue
        col = column_index_from_string(
            "".join(ch for ch in coord if ch.isalpha())
        )
        row = int("".join(ch for ch in coord if ch.isdigit()))
        if col <= INPUT_MAX_COL and row <= INPUT_MAX_ROW:
            picked.add((sheet, coord))
    return picked


class OptimizingEmitter(Emitter):
    chunked = True

    def __init__(self, g: WorkbookGraph, inputs: set[tuple[str, str]] | None = None):
        super().__init__(g, default_inputs(g) if inputs is None else inputs)
        self._throws = 0          # 动态 VLOOKUP 计数,IFERROR 剥除的依据
        self.stat_vl = self.stat_vl_kept = 0
        self.stat_xl = self.stat_xl_kept = 0
        self.stat_iferr_dropped = 0

    # -- 静态求值

    def static_value(self, n: Node, sheet: str):
        """字面量 / 指向非输入字面量格的单格引用 -> Python 值;
        _EMPTY = 静态空格;None = 无法静态求值。"""
        if isinstance(n, Num):
            return float(n.text)
        if isinstance(n, Str):
            return n.value
        if isinstance(n, Bool):
            return n.value
        if isinstance(n, Ref) and n.end is None:
            s = n.sheet or sheet
            key = (s, f"{n.start.col}{n.start.row}")
            if key in self.inputs or key in self.g.nodes:
                return None
            if key in self.g.values:
                v = self.g.values[key]
                return float(v) if isinstance(v, (int, float)) and not isinstance(v, bool) else v
            return _EMPTY
        return None

    @staticmethod
    def _match(cell_v, key) -> bool:
        """镜像运行时 VL/XL 的命中判定(文本不区分大小写)。"""
        if isinstance(cell_v, str):
            return isinstance(key, str) and cell_v.lower() == key.lower()
        return cell_v == key

    def _static_column(self, span, want_col: int) -> Optional[list]:
        """区间第 want_col 列(0-based)逐行取静态值;含公式格或输入格则放弃。"""
        sheet, r1, c1, r2, c2 = span
        if c1 + want_col > c2:
            return None
        out = []
        for r in range(r1, r2 + 1):
            key = (sheet, f"{get_column_letter(c1 + want_col)}{r}")
            if key in self.g.nodes or key in self.inputs:
                return None
            out.append(self.g.values.get(key, _EMPTY))
        return out

    def _target_js(self, span, row_off: int, col_off: int) -> str:
        sheet, r1, c1, r2, c2 = span
        if c1 + col_off > c2:
            return "null"
        key = (sheet, f"{get_column_letter(c1 + col_off)}{r1 + row_off}")
        i = self.slot.get(key, -1)
        return "null" if i < 0 else f"x[{i}]"

    # -- 去虚化

    def _devirt_vlookup(self, n: Call, sheet: str) -> Optional[str]:
        key = self.static_value(n.args[0], sheet)
        if key is None or key is _EMPTY:
            return None
        col = self.static_value(n.args[2], sheet)
        if not isinstance(col, float):
            return None
        if not isinstance(n.args[1], Ref) or n.args[1].end is None:
            return None
        span = resolve_ref(n.args[1], sheet, self.g.sheet_extent)
        first = self._static_column(span, 0)
        if first is None:
            return None
        for off, v in enumerate(first):
            if v is not _EMPTY and self._match(v, key):
                return self._target_js(span, off, int(col) - 1)
        return None  # 静态未命中:保守回退动态(实为 #N/A,语料中不出现)

    def _devirt_xlookup(self, n: Call, sheet: str) -> Optional[str]:
        if not isinstance(n.args[1], Ref) or n.args[1].end is None:
            return None
        if not isinstance(n.args[2], Ref) or n.args[2].end is None:
            return None
        ifnf = self.static_value(n.args[3], sheet)
        if ifnf is None:
            return None
        la = resolve_ref(n.args[1], sheet, self.g.sheet_extent)
        ra = resolve_ref(n.args[2], sheet, self.g.sheet_extent)
        lookup = self._static_column(la, 0)
        if lookup is None:
            return None

        # 查找键:标量或全静态区间
        vals = []
        if isinstance(n.args[0], Ref) and n.args[0].end is not None:
            vspan = resolve_ref(n.args[0], sheet, self.g.sheet_extent)
            vs, vr1, vc1, vr2, vc2 = vspan
            for r in range(vr1, vr2 + 1):
                for c in range(vc1, vc2 + 1):
                    key = (vs, f"{get_column_letter(c)}{r}")
                    if key in self.g.nodes or key in self.inputs:
                        return None
                    vals.append(self.g.values.get(key, _EMPTY))
        else:
            v = self.static_value(n.args[0], sheet)
            if v is None:
                return None
            vals.append(v)

        if not isinstance(ifnf, float):
            return None  # 语料中恒为数字 0,其余形态不折叠
        ifnf_js = repr(ifnf)
        parts = []
        for kv in vals:
            hit = None
            if kv is not _EMPTY:
                for off, lv in enumerate(lookup):
                    if lv is not _EMPTY and self._match(lv, kv):
                        hit = off
                        break
            parts.append(self._target_js(ra, hit, 0) if hit is not None else ifnf_js)
        if len(parts) == 1:
            return parts[0]
        return "[" + ",".join(parts) + "]"

    # -- 编译改写

    def js(self, n: Node, sheet: str) -> str:
        if isinstance(n, Call):
            name = n.name.upper()
            if name == "VLOOKUP":
                self.stat_vl += 1
                folded = self._devirt_vlookup(n, sheet)
                if folded is not None:
                    return folded
                self.stat_vl_kept += 1
                self._throws += 1
                return super().js(n, sheet)
            if name == "_XLFN.XLOOKUP":
                self.stat_xl += 1
                folded = self._devirt_xlookup(n, sheet)
                if folded is not None:
                    return folded
                self.stat_xl_kept += 1
                return super().js(n, sheet)
            if name == "IFERROR":
                mark = self._throws
                a0 = self.js(n.args[0], sheet)
                if self._throws == mark:
                    self.stat_iferr_dropped += 1
                    return f"({a0})"
                return f"IFERR(()=>({a0}),{self.js(n.args[1], sheet)})"
        return super().js(n, sheet)

    def report(self) -> str:
        return (
            f"VLOOKUP {self.stat_vl} 折叠 {self.stat_vl - self.stat_vl_kept}, "
            f"XLOOKUP {self.stat_xl} 折叠 {self.stat_xl - self.stat_xl_kept}, "
            f"IFERROR 剥除 {self.stat_iferr_dropped}"
        )


def emit_optimized(path: Path) -> tuple[str, str]:
    em = OptimizingEmitter(load_graph(path))
    src = em.emit()
    return src, em.report()


if __name__ == "__main__":
    src_path = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else src_path.with_suffix(".opt.js")
    src, report = emit_optimized(src_path)
    out.write_text(src, encoding="utf-8")
    print(f"{src_path.name} -> {out} ({out.stat().st_size / 1024:.0f} KB)\n  {report}")
