"""工作簿依赖图:解析全部公式格,解析引用,建图并拓扑排序。

图模型:
- 节点 = 带公式的单元格,键为 (表名, 坐标)。
- 字面量格(数字/文本)不是节点,是常量叶子,求值时直接取值。
- 边 = 公式格 A 引用了公式格 B(A 依赖 B)。对区间引用,展开为区间内
  全部非空格,其中公式格计入依赖边。
- 整列引用($A:$ZZ、L:L)裁剪到目标表的实际使用范围(max_row/max_column),
  这是 anatomy 第 4 节点名的 codegen 义务。

拓扑排序用 Kahn 算法,同时产出分层(第 n 层只依赖前 n-1 层),层数即
求值深度。检出环则抛 CycleError —— 上游表已证实无迭代计算,出环
一定是解析缺陷。

作为脚本运行时对 upstream/ 全部 xlsx 建图并输出统计,任一文件出环
即退出码 1。
"""

from __future__ import annotations

import sys
from collections import deque
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterator, Optional

import openpyxl
from openpyxl.utils import column_index_from_string, get_column_letter
from openpyxl.worksheet.formula import ArrayFormula

from formula_parser import (
    Bin, Bool, Call, CellPart, Num, Paren, Pct, Ref, Str, Un,
    Node, parse_formula,
)

Key = tuple[str, str]  # (表名, "B22")


class CycleError(ValueError):
    pass


@dataclass
class CellNode:
    sheet: str
    coord: str
    formula: str          # 原文(含 '='),数组公式取其 text
    is_array: bool
    ast: Node
    cached: object        # data_only=True 的缓存值,金标准
    deps: set[Key] = field(default_factory=set)


@dataclass
class WorkbookGraph:
    path: Path
    nodes: dict[Key, CellNode]
    values: dict[Key, object]         # 字面量格
    sheet_extent: dict[str, tuple[int, int]]  # 表名 -> (max_row, max_col)
    topo: list[Key] = field(default_factory=list)
    layers: list[list[Key]] = field(default_factory=list)

    @property
    def edge_count(self) -> int:
        return sum(len(n.deps) for n in self.nodes.values())


def iter_refs(node: Node) -> Iterator[Ref]:
    """深度优先枚举 AST 里的全部 Ref。"""
    if isinstance(node, Ref):
        yield node
    elif isinstance(node, Call):
        for a in node.args:
            yield from iter_refs(a)
    elif isinstance(node, Bin):
        yield from iter_refs(node.left)
        yield from iter_refs(node.right)
    elif isinstance(node, (Un, Pct)):
        yield from iter_refs(node.operand)
    elif isinstance(node, Paren):
        yield from iter_refs(node.inner)
    # Num/Str/Bool 无引用


def resolve_ref(
    ref: Ref, current_sheet: str, extent: dict[str, tuple[int, int]]
) -> tuple[str, int, int, int, int]:
    """Ref -> (表名, r1, c1, r2, c2),1-based 闭区间,已裁剪到使用范围。"""
    sheet = ref.sheet or current_sheet
    if sheet not in extent:
        raise CycleError(f"引用了不存在的表 {sheet!r}")
    max_row, max_col = extent[sheet]

    def part(p: CellPart) -> tuple[Optional[int], Optional[int]]:
        c = column_index_from_string(p.col) if p.col is not None else None
        return p.row, c

    r1, c1 = part(ref.start)
    if ref.end is None:
        r2, c2 = r1, c1
    else:
        r2, c2 = part(ref.end)

    # 整列引用:行端为 None -> 1..max_row
    if r1 is None:
        r1 = 1
    if r2 is None:
        r2 = max_row
    # 行区间(语料中不存在,防御):列端为 None -> 1..max_col
    if c1 is None:
        c1 = 1
    if c2 is None:
        c2 = max_col

    r1, r2 = min(r1, r2), max(r1, r2)
    c1, c2 = min(c1, c2), max(c1, c2)
    return sheet, r1, c1, min(r2, max_row), min(c2, max_col)


def load_graph(path: Path) -> WorkbookGraph:
    wbf = openpyxl.load_workbook(path, data_only=False)
    wbv = openpyxl.load_workbook(path, data_only=True)

    nodes: dict[Key, CellNode] = {}
    values: dict[Key, object] = {}
    extent: dict[str, tuple[int, int]] = {}

    for ws in wbf.worksheets:
        extent[ws.title] = (ws.max_row, ws.max_column)
        wsv = wbv[ws.title]
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if v is None:
                    continue
                key = (ws.title, c.coordinate)
                if isinstance(v, ArrayFormula):
                    text = v.text
                elif isinstance(v, str) and v.startswith("="):
                    text = v
                else:
                    values[key] = v
                    continue
                nodes[key] = CellNode(
                    sheet=ws.title, coord=c.coordinate,
                    formula=text, is_array=isinstance(v, ArrayFormula),
                    ast=parse_formula(text),
                    cached=wsv[c.coordinate].value,
                )

    g = WorkbookGraph(path=path, nodes=nodes, values=values, sheet_extent=extent)

    # 依赖解析。区间 -> 公式格依赖集 做记忆化:同列的 135 行轴公式
    # 反复引用同一批 A:ZZ 整列区间,不缓存会白展开几百万次。
    range_cache: dict[tuple[str, int, int, int, int], frozenset[Key]] = {}

    def deps_of_span(span: tuple[str, int, int, int, int]) -> frozenset[Key]:
        if span in range_cache:
            return range_cache[span]
        sheet, r1, c1, r2, c2 = span
        found = frozenset(
            key
            for r in range(r1, r2 + 1)
            for co in range(c1, c2 + 1)
            if (key := (sheet, f"{get_column_letter(co)}{r}")) in nodes
        )
        range_cache[span] = found
        return found

    for key, n in nodes.items():
        for ref in iter_refs(n.ast):
            n.deps |= deps_of_span(resolve_ref(ref, n.sheet, extent))
        n.deps.discard(key)  # 语料中不存在自引用,防御

    # Kahn 拓扑 + 分层
    indeg = {k: len(n.deps) for k, n in nodes.items()}
    dependents: dict[Key, list[Key]] = {k: [] for k in nodes}
    for k, n in nodes.items():
        for d in n.deps:
            dependents[d].append(k)

    frontier = deque(sorted(k for k, d in indeg.items() if d == 0))
    while frontier:
        layer = sorted(frontier)
        frontier.clear()
        g.layers.append(layer)
        g.topo.extend(layer)
        for k in layer:
            for m in dependents[k]:
                indeg[m] -= 1
                if indeg[m] == 0:
                    frontier.append(m)

    if len(g.topo) != len(nodes):
        stuck = sorted(k for k, d in indeg.items() if d > 0)[:10]
        raise CycleError(f"{path.name}: 依赖成环,涉及 {len(nodes) - len(g.topo)} 格,如 {stuck}")

    return g


def _selfcheck(g: WorkbookGraph) -> None:
    pos = {k: i for i, k in enumerate(g.topo)}
    for k, n in g.nodes.items():
        for d in n.deps:
            assert pos[d] < pos[k], f"拓扑序违例: {d} 应先于 {k}"


def main() -> None:
    target = Path(sys.argv[1] if len(sys.argv) > 1 else Path(__file__).parents[2] / "upstream")
    files = [target] if target.is_file() else sorted(target.glob("*.xlsx"))
    if not files:
        sys.exit(f"{target} 下没有 xlsx")

    failed = False
    for f in files:
        try:
            g = load_graph(f)
            _selfcheck(g)
            print(
                f"{f.name}: 公式格 {len(g.nodes)}, 字面量格 {len(g.values)}, "
                f"边 {g.edge_count}, 层数 {len(g.layers)}, "
                f"最大层 {max(map(len, g.layers))}"
            )
        except CycleError as e:
            failed = True
            print(f"FAIL {e}")
    if failed:
        sys.exit(1)


if __name__ == "__main__":
    main()
