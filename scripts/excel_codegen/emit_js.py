"""JS 生成器:依赖图 → 自包含可执行 JS。

生成物结构:
  1. 运行时前导(RUNTIME):Excel 语义的 9 个函数 + 类型强转。
     实测语料形态(见 git log / anatomy):IF 恒 3 参、VLOOKUP 恒
     精确匹配、XLOOKUP 恒 4 参且查找键可为区间(数组语义仅此一处)、
     比较符只有 = > < >=、无字符串拼接无乘方。
  2. 单元格槽数组 x:每个非空格一个下标,字面量格直接赋值。
  3. 区间常量:{c: 列数, i: Int32 下标数组(行主序, -1=空格)},全簿去重。
  4. F:拓扑序的 (下标, 求值闭包) 列表,逐条 try 执行,错误落入
     {__err} 而不中断(缓存值已证实零错误,任何 __err 都是生成缺陷)。
  5. 尾部 console.log(JSON) 输出全部公式格结果,供金标准比对。

Excel 语义要点:空格在算术中为 0、在与字符串比较中为 "";文本比较
不区分大小写;SUM 忽略文本与空;VLOOKUP 未命中抛 #N/A(IFERROR 捕获);
XLOOKUP 未命中返回第 4 参。
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from formula_parser import (
    Bin, Bool, Call, Num, Paren, Pct, Ref, Str, Un, Node,
)
from workbook_graph import WorkbookGraph, load_graph, resolve_ref

RUNTIME = r"""
"use strict";
const ERR = (c) => { const e = new Error(c); e.excel = c; throw e; };
const N = (v) => typeof v === "number" ? v : v == null ? 0
  : v === true ? 1 : v === false ? 0 : ERR("#VALUE!");
const B = (v) => typeof v === "boolean" ? v
  : typeof v === "number" ? v !== 0 : v == null ? false : ERR("#VALUE!");
const EQ = (a, b) => {
  if (a == null) a = typeof b === "string" ? "" : 0;
  if (b == null) b = typeof a === "string" ? "" : 0;
  if (typeof a === "string" && typeof b === "string")
    return a.toLowerCase() === b.toLowerCase();
  return a === b;
};
const GT = (a, b) => N(a) > N(b);
const GE = (a, b) => N(a) >= N(b);
const LT = (a, b) => N(a) < N(b);
const LE = (a, b) => N(a) <= N(b);
function SUM(...as) {
  let s = 0;
  for (const a of as) {
    if (a && a.i) {
      const i = a.i;
      for (let k = 0; k < i.length; k++) {
        const j = i[k];
        if (j >= 0) { const v = x[j]; if (typeof v === "number") s += v; }
      }
    } else if (Array.isArray(a)) {
      for (const v of a) if (typeof v === "number") s += v;
    } else if (typeof a === "number") s += a;
    else if (a === true) s += 1;
  }
  return s;
}
function VL(key, rg, col) {
  col = N(col) | 0;
  const c = rg.c, i = rg.i, rows = i.length / c;
  const ks = typeof key === "string" ? key.toLowerCase() : key;
  for (let r = 0; r < rows; r++) {
    const j = i[r * c];
    if (j < 0) continue;
    const v = x[j];
    const hit = typeof v === "string"
      ? (typeof ks === "string" && v.toLowerCase() === ks)
      : v === key;
    if (hit) { const t = i[r * c + col - 1]; return t < 0 ? null : x[t]; }
  }
  ERR("#N/A");
}
function XL(vals, la, ra, ifnf) {
  const one = (kv) => {
    const ks = typeof kv === "string" ? kv.toLowerCase() : kv;
    const c = la.c, i = la.i, rows = i.length / c;
    for (let r = 0; r < rows; r++) {
      const j = i[r * c];
      if (j < 0) continue;
      const v = x[j];
      const hit = typeof v === "string"
        ? (typeof ks === "string" && v.toLowerCase() === ks)
        : v === kv;
      if (hit) { const t = ra.i[r * ra.c]; return t < 0 ? null : x[t]; }
    }
    return ifnf;
  };
  if (vals && vals.i) {
    const out = [], i = vals.i;
    for (let k = 0; k < i.length; k++) {
      const j = i[k];
      out.push(one(j < 0 ? null : x[j]));
    }
    return out;
  }
  return one(vals);
}
const IFERR = (f, alt) => { try { return f(); } catch (e) { if (e && e.excel) return alt; throw e; } };
const ROUNDUP = (v, d) => {
  const m = 10 ** (N(d) | 0), u = N(v);
  return u >= 0 ? Math.ceil(u * m) / m : -Math.ceil(-u * m) / m;
};
"""

_CMP = {"=": "EQ", ">": "GT", ">=": "GE", "<": "LT", "<=": "LE"}


class Emitter:
    chunked = False  # True: 分块直线代码(OptimizingEmitter 用),False: 闭包表 + 逐格 try

    def __init__(self, g: WorkbookGraph, inputs: set[tuple[str, str]] | None = None):
        self.g = g
        # 输入格:运行时可被覆写的字面量格(优化器据此决定什么不可折叠)
        self.inputs = inputs or set()
        # 全部非空格 -> 槽下标;声明为输入的空格也分配槽位(如鸣金虹
        # 半接线的 期望!C21),否则引用处会被编译成写不进值的 null
        self.slot: dict[tuple[str, str], int] = {}
        for key in list(g.values) + list(g.nodes) + sorted(self.inputs):
            if key not in self.slot:
                self.slot[key] = len(self.slot)
        self.ranges: dict[tuple[str, int, int, int, int], str] = {}
        self.range_defs: list[str] = []

    # -- 区间常量

    def range_const(self, span: tuple[str, int, int, int, int]) -> str:
        if span in self.ranges:
            return self.ranges[span]
        name = f"R{len(self.ranges)}"
        self.ranges[span] = name
        sheet, r1, c1, r2, c2 = span
        from openpyxl.utils import get_column_letter
        idx = [
            self.slot.get((sheet, f"{get_column_letter(c)}{r}"), -1)
            for r in range(r1, r2 + 1)
            for c in range(c1, c2 + 1)
        ]
        self.range_defs.append(
            f"const {name}={{c:{c2 - c1 + 1},i:Int32Array.from({json.dumps(idx)})}};"
        )
        return name

    # -- 表达式编译

    def _numeric(self, n: Node) -> bool:
        """静态可证结果必为数值(JS number)的节点,算术处可跳过 N() 强转。"""
        if isinstance(n, (Num, Pct, Un)):
            return True
        if isinstance(n, Bin):
            return n.op in "+-*/^"
        if isinstance(n, Paren):
            return self._numeric(n.inner)
        if isinstance(n, Call):
            return n.name.upper() in ("SUM", "MIN", "MAX", "ROUNDUP")
        return False

    def _njs(self, n: Node, sheet: str) -> str:
        s = self.js(n, sheet)
        return s if self._numeric(n) else f"N({s})"

    def js(self, n: Node, sheet: str) -> str:
        if isinstance(n, Num):
            return n.text
        if isinstance(n, Str):
            return json.dumps(n.value, ensure_ascii=False)
        if isinstance(n, Bool):
            return "true" if n.value else "false"
        if isinstance(n, Paren):
            return f"({self.js(n.inner, sheet)})"
        if isinstance(n, Un):
            return f"({n.op}{self._njs(n.operand, sheet)})"
        if isinstance(n, Pct):
            return f"({self._njs(n.operand, sheet)}/100)"
        if isinstance(n, Ref):
            span = resolve_ref(n, sheet, self.g.sheet_extent)
            _, r1, c1, r2, c2 = span
            if n.end is None and r1 == r2 and c1 == c2:
                i = self.slot.get((span[0], f"{_col(c1)}{r1}"), -1)
                return "null" if i < 0 else f"x[{i}]"
            return self.range_const(span)
        if isinstance(n, Bin):
            # 注意:每个子树只编译一次,先编译再按需包 N() —— 若在此处
            # 对子树重复调用 js(),深层嵌套公式会指数爆炸
            if n.op in ("+", "-", "*", "/", "^"):
                l, r = self._njs(n.left, sheet), self._njs(n.right, sheet)
                op = "**" if n.op == "^" else n.op
                return f"({l}{op}{r})"
            l, r = self.js(n.left, sheet), self.js(n.right, sheet)
            if n.op in _CMP:
                return f"{_CMP[n.op]}({l},{r})"
            if n.op == "<>":
                return f"(!EQ({l},{r}))"
            if n.op == "&":
                return f"(String({l})+String({r}))"
            raise ValueError(f"未知运算符 {n.op}")
        if isinstance(n, Call):
            # 每个实参只编译一次(不预编译整表:MIN/MAX 走 _njs,重复
            # 编译会导致子树被访问两次,统计失真且拖慢生成)
            name = n.name.upper()
            def a(k: int) -> str:
                return self.js(n.args[k], sheet)
            if name == "IF":
                return f"(B({a(0)})?{a(1)}:{a(2)})"
            if name == "OR":
                return "(" + "||".join(f"B({self.js(arg, sheet)})" for arg in n.args) + ")"
            if name == "SUM":
                return f"SUM({','.join(self.js(arg, sheet) for arg in n.args)})"
            if name == "MIN":
                return f"Math.min({','.join(self._njs(arg, sheet) for arg in n.args)})"
            if name == "MAX":
                return f"Math.max({','.join(self._njs(arg, sheet) for arg in n.args)})"
            if name == "VLOOKUP":
                return f"VL({a(0)},{a(1)},{a(2)})"
            if name == "_XLFN.XLOOKUP":
                return f"XL({a(0)},{a(1)},{a(2)},{a(3)})"
            if name == "IFERROR":
                return f"IFERR(()=>({a(0)}),{a(1)})"
            if name == "ROUNDUP":
                return f"ROUNDUP({a(0)},{a(1)})"
            raise ValueError(f"未实现的函数 {n.name}")
        raise ValueError(f"未知节点 {n!r}")

    # -- 整体生成

    def literal_lines(self) -> list[str]:
        out = []
        for key, v in self.g.values.items():
            i = self.slot[key]
            if isinstance(v, bool):
                out.append(f"x[{i}]={'true' if v else 'false'};")
            elif isinstance(v, (int, float)):
                out.append(f"x[{i}]={v!r};")
            elif isinstance(v, str):
                out.append(f"x[{i}]={json.dumps(v, ensure_ascii=False)};")
            else:  # 日期等,语料中仅更新日志出现且无公式引用
                out.append(f"x[{i}]={json.dumps(str(v), ensure_ascii=False)};")
        return out

    def topo_statements(self) -> list[str]:
        return [
            f"x[{self.slot[key]}]={self.js(self.g.nodes[key].ast, self.g.nodes[key].sheet)};"
            for key in self.g.topo
        ]

    def chunk_lines(self, stmts: list[str], size: int = 500) -> list[str]:
        chunks = [stmts[i:i + size] for i in range(0, len(stmts), size)]
        out = [
            f"function C{ci}(){{\n" + "\n".join(chunk) + "\n}"
            for ci, chunk in enumerate(chunks)
        ]
        out.append(
            "function EVAL(){" + "".join(f"C{ci}();" for ci in range(len(chunks))) + "}"
        )
        return out

    def emit(self) -> str:
        g = self.g
        lines = [RUNTIME, f"const x=new Array({len(self.slot)}).fill(null);"]
        lines.extend(self.literal_lines())

        # 输入覆写钩子:node argv[2] 传 JSON {"表!坐标": 值} 可改写输入格,
        # 供等价性模糊测试与将来的求解器注入面板数据
        slot_map = {f"{s}!{c}": self.slot[(s, c)] for s, c in sorted(self.inputs)}
        lines.append(f"const SLOT={json.dumps(slot_map, ensure_ascii=False)};")
        lines.append(
            'const OVR=(typeof process!=="undefined"&&process.argv[2])'
            '?JSON.parse(process.argv[2]):{};'
            "for(const k in OVR){if(k in SLOT)x[SLOT[k]]=OVR[k];}"
        )

        if self.chunked:
            # 直线代码分块:避免 12k 闭包的逐个调用开销,也避免单个
            # 超大函数体让 V8 放弃优化。整块 try:金标准已证零错误,
            # 任何异常都是生成缺陷,响亮失败即可。
            body = self.chunk_lines(self.topo_statements())
            body.append('try{EVAL();}catch(e){console.error("EVAL 异常:",e);process.exit(3);}')
        else:
            body = []
            for key in g.topo:
                n = g.nodes[key]
                body.append(f"F.push([{self.slot[key]},()=>{self.js(n.ast, n.sheet)}]);")

        lines.extend(self.range_defs)
        if not self.chunked:
            lines.append("const F=[];")
        lines.extend(body)
        if not self.chunked:
            lines.append(
                "for(const [i,fn] of F){try{x[i]=fn();}catch(e){x[i]={__err:String(e.excel||e)};}}"
            )
        keys = {f"{s}!{c}": self.slot[(s, c)] for s, c in g.nodes}
        lines.append(f"const K={json.dumps(keys, ensure_ascii=False)};")
        lines.append(
            "const out={};for(const k in K)out[k]=x[K[k]];"
            "console.log(JSON.stringify(out));"
        )
        return "\n".join(lines)


def _col(c: int) -> str:
    from openpyxl.utils import get_column_letter
    return get_column_letter(c)


def emit_workbook(path: Path) -> str:
    return Emitter(load_graph(path)).emit()


if __name__ == "__main__":
    src = Path(sys.argv[1])
    out = Path(sys.argv[2]) if len(sys.argv) > 2 else src.with_suffix(".gen.js")
    out.write_text(emit_workbook(src), encoding="utf-8")
    print(f"{src.name} -> {out} ({out.stat().st_size / 1024:.0f} KB)")
